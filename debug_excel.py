import os
import openpyxl
from knowledge_base.ingest import ExcelLoader

# Create a dummy Excel file
filename = "debug_test.xlsx"
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sheet1"
ws['A1'] = "Test Header"
ws['B1'] = "Test Value"
ws['A2'] = "Row 2 Col A"
ws['B2'] = "Row 2 Col B"
wb.save(filename)

print(f"Created {filename}")

# Try to load it
try:
    print("Attempting to load with ExcelLoader...")
    loader = ExcelLoader(filename)
    docs = loader.load()
    print(f"Success! Loaded {len(docs)} documents.")
    for doc in docs:
        print("--- Content ---")
        print(doc.page_content)
        print("--- Metadata ---")
        print(doc.metadata)
except Exception as e:
    print("!!! Failed to load !!!")
    print(e)
    import traceback
    traceback.print_exc()
